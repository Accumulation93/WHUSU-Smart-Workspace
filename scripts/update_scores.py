#!/usr/bin/env python3
"""
Update/insert 外联客观评分员's scoring records from ~/外联评分修改更新.xlsx Sheet3.

For each 学员 (student) in Sheet3:
  - If a score_record exists (rule 7a15c9e5, scorer=外联客观评分员): UPDATE the score_answer
  - If not: INSERT a new score_record + score_answer

The score value is rounded to 2 decimal places.
"""

import os
import sys
import hashlib
import pymysql
import openpyxl

# ── Constants (from approved plan) ──────────────────────────────────────
ORG_ID     = "3ea5f6fa5f23758773267fc1d2803673bf71f96dbde6c59fdb543f8e3db5a1c6"
ACTIVITY_ID = "115c9589a5b1cfce80fc27a4697e8eaf18f1d31631b3827483749aad560333ca"
SCORER_ID  = "467d0e7c65914ec488bed4215cf537da34f94b68f87a5e88f442663f92a78e29"
RULE_ID    = "7a15c9e5e25ed444b080ae25b2ce9781d70e99b89bcd17d519507465d09f7721"
TEMPLATE_SIG = "4ed920d37165e01e1102a949e2626dad8a6a9a345a6c5a3b96a20950191dee63[1|weighted_average|0|0]"
IDENTITY_ID = "3d510268afc113c3a1e1c7223a5f9a7b5870e874041d590f6b9ea02443a62632"

DB_CONFIG = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "redsu",
    "password": "e9cadd9f07e3f76e8a518956e1062ee8",
    "database": "redsu_scoring",
    "charset": "utf8mb4",
}

XLSX_PATH = os.path.expanduser("~/外联评分修改更新.xlsx")


def generate_id():
    """Mimic server's generateId(): crypto.randomBytes(32).toString('hex')"""
    return os.urandom(32).hex()


def read_sheet3(path):
    """Parse Sheet3, return list of {name, score} with score rounded to 2dp."""
    wb = openpyxl.load_workbook(path)
    ws = wb["Sheet3"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name = row[0]
        score = row[2]  # Column C = 总分
        if name is not None and score is not None:
            rows.append({
                "name": str(name).strip(),
                "score": round(float(score), 2),
            })
    return rows


def main():
    # ── 1. Read Excel ───────────────────────────────────────────────────
    print("Reading Sheet3 from", XLSX_PATH)
    entries = read_sheet3(XLSX_PATH)
    print(f"  Found {len(entries)} entries")
    for i, e in enumerate(entries):
        print(f"  [{i+1:2d}] {e['name']:　<6s} → {e['score']:.2f}")

    # ── 2. Connect to DB ────────────────────────────────────────────────
    print("\nConnecting to MySQL...")
    conn = pymysql.connect(**DB_CONFIG)
    cursor = conn.cursor()

    # ── 3. Match names to hr_info ───────────────────────────────────────
    print("Matching names to hr_info...")
    name_to_hr = {}
    hr_to_name = {}
    not_found = []

    cursor.execute(
        "SELECT id, name FROM hr_info WHERE identity_id = %s AND org_id = %s",
        (IDENTITY_ID, ORG_ID)
    )
    for hr_id, hr_name in cursor.fetchall():
        name_to_hr[hr_name.strip()] = hr_id
        hr_to_name[hr_id] = hr_name.strip()

    for entry in entries:
        if entry["name"] in name_to_hr:
            entry["target_id"] = name_to_hr[entry["name"]]
        else:
            not_found.append(entry["name"])

    if not_found:
        print(f"\n  ⚠ WARNING: {len(not_found)} names NOT found in hr_info (identity=学员):")
        for n in not_found:
            print(f"    - {n}")

    matched = [e for e in entries if "target_id" in e]
    print(f"  Matched: {len(matched)}, Not found: {len(not_found)}")

    # ── 4. Find existing records ────────────────────────────────────────
    print("\nChecking existing records...")
    target_ids = [e["target_id"] for e in matched]
    placeholders = ",".join(["%s"] * len(target_ids))

    cursor.execute(
        f"""SELECT sr.id, sr.target_id, sa.score
            FROM score_records sr
            JOIN score_answers sa ON sa.record_id = sr.id
            WHERE sr.activity_id = %s
              AND sr.rule_id = %s
              AND sr.scorer_id = %s
              AND sr.org_id = %s
              AND sr.target_id IN ({placeholders})""",
        (ACTIVITY_ID, RULE_ID, SCORER_ID, ORG_ID, *target_ids)
    )
    existing = {}
    for rec_id, target_id, score in cursor.fetchall():
        existing[target_id] = {"record_id": rec_id, "old_score": float(score)}

    to_update = []
    to_insert = []
    for entry in matched:
        tid = entry["target_id"]
        if tid in existing:
            to_update.append({**entry, **existing[tid]})
        else:
            to_insert.append(entry)

    print(f"  Will UPDATE: {len(to_update)} records")
    print(f"  Will INSERT:  {len(to_insert)} records")

    # ── 5. Execute updates/inserts ──────────────────────────────────────
    print("\nExecuting changes...")
    updated_count = 0
    inserted_count = 0

    for entry in to_update:
        cursor.execute(
            "UPDATE score_answers SET score = %s WHERE record_id = %s AND org_id = %s",
            (entry["score"], entry["record_id"], ORG_ID)
        )
        updated_count += 1
        name = entry["name"]
        old = entry["old_score"]
        new = entry["score"]
        print(f"  UPDATE {name:　<6s}: {old:.2f} → {new:.2f}")

    for entry in to_insert:
        new_record_id = generate_id()
        new_answer_id = generate_id()
        tid = entry["target_id"]
        score = entry["score"]

        cursor.execute(
            """INSERT INTO score_records
               (id, activity_id, rule_id, scorer_id, target_id,
                template_config_signature, org_id)
               VALUES (%s, %s, %s, %s, %s, %s, %s)""",
            (new_record_id, ACTIVITY_ID, RULE_ID, SCORER_ID, tid,
             TEMPLATE_SIG, ORG_ID)
        )
        cursor.execute(
            """INSERT INTO score_answers
               (id, record_id, question_index, score, org_id)
               VALUES (%s, %s, 1, %s, %s)""",
            (new_answer_id, new_record_id, score, ORG_ID)
        )
        inserted_count += 1
        print(f"  INSERT {entry['name']:　<6s}: → {score:.2f}")

    conn.commit()
    print(f"\n  ✓ Committed: {updated_count} updated, {inserted_count} inserted")

    # ── 6. Verify ───────────────────────────────────────────────────────
    print("\n─── Verification ───")
    cursor.execute(
        f"""SELECT h.name, sa.score
            FROM score_records sr
            JOIN hr_info h ON h.id = sr.target_id
            JOIN score_answers sa ON sa.record_id = sr.id
            WHERE sr.activity_id = %s
              AND sr.rule_id = %s
              AND sr.scorer_id = %s
              AND sr.org_id = %s
            ORDER BY h.name""",
        (ACTIVITY_ID, RULE_ID, SCORER_ID, ORG_ID)
    )
    db_after = {}
    for name, score in cursor.fetchall():
        db_after[name.strip()] = float(score)

    print(f"  Records in DB: {len(db_after)}")
    discrepancies = []
    for entry in matched:
        db_score = db_after.get(entry["name"])
        expected = entry["score"]
        if db_score is None:
            discrepancies.append(f"  MISSING: {entry['name']}")
        elif abs(db_score - expected) > 0.01:
            discrepancies.append(
                f"  MISMATCH: {entry['name']} expected {expected:.2f}, got {db_score:.2f}"
            )

    if discrepancies:
        print(f"  ⚠ {len(discrepancies)} issues found:")
        for d in discrepancies:
            print(d)
    else:
        print(f"  ✓ All {len(matched)} entries match expected scores")

    # Summary
    print(f"\n─── Summary ───")
    print(f"  Excel entries:   {len(entries)}")
    print(f"  Names not in DB: {len(not_found)}")
    print(f"  Updated:         {updated_count}")
    print(f"  Inserted:        {inserted_count}")
    print(f"  Total in DB now: {len(db_after)}")

    cursor.close()
    conn.close()


if __name__ == "__main__":
    main()
