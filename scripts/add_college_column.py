#!/usr/bin/env python3
"""
Read ~/第二学期名单.xlsx, match each person to their 学院 (college) from
the database's extended HR profile, and insert a 学院 column between
B (学号) and C (部门).

Output: ~/第二学期名单_含学院.xlsx
"""

import os
import pymysql
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from collections import OrderedDict

# ── Config ──────────────────────────────────────────────────────────────
XLSX_IN  = os.path.expanduser("~/第二学期名单.xlsx")
XLSX_OUT = os.path.expanduser("~/第二学期名单_含学院.xlsx")

DB_CONFIG = {
    "host": "127.0.0.1",
    "port": 3306,
    "user": "redsu",
    "password": "e9cadd9f07e3f76e8a518956e1062ee8",
    "database": "redsu_scoring",
    "charset": "utf8mb4",
}

COLLEGE_FIELD_ID = "profile_field_1780049686656_524"
ORG_ID = "3ea5f6fa5f23758773267fc1d2803673bf71f96dbde6c59fdb543f8e3db5a1c6"


def get_college_map(conn, names):
    """Return {name: college} for the given list of names."""
    if not names:
        return {}

    placeholders = ",".join(["%s"] * len(names))
    sql = f"""
        SELECT h.name, pv.field_value
        FROM hr_info h
        LEFT JOIN hr_profile_records pr ON pr.hr_id = h.id
        LEFT JOIN hr_profile_record_values pv ON pv.record_id = pr.id
            AND pv.field_id = %s
        WHERE h.name IN ({placeholders})
          AND h.org_id = %s
    """
    params = [COLLEGE_FIELD_ID] + list(names) + [ORG_ID]

    cursor = conn.cursor()
    cursor.execute(sql, params)
    result = {}
    for name, college in cursor.fetchall():
        result[name.strip()] = college.strip() if college else ""
    cursor.close()
    return result


def main():
    # ── 1. Read source xlsx ─────────────────────────────────────────────
    print(f"Reading: {XLSX_IN}")
    wb = openpyxl.load_workbook(XLSX_IN)
    ws = wb.active
    print(f"  Sheet: {ws.title}, {ws.max_row} rows × {ws.max_column} cols")

    # Read headers and data
    headers = [cell.value for cell in ws[1]]
    print(f"  Headers: {headers}")

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        rows.append(list(row))

    names_in_file = [r[0].strip() if r[0] else "" for r in rows]
    print(f"  Data rows: {len(rows)}")

    # ── 2. Query college from DB ────────────────────────────────────────
    print("\nQuerying college info from database...")
    conn = pymysql.connect(**DB_CONFIG)
    college_map = get_college_map(conn, names_in_file)
    conn.close()

    found = sum(1 for v in college_map.values() if v)
    empty = sum(1 for v in college_map.values() if not v)
    print(f"  Found college: {found}, Empty: {empty}")

    # Show missing
    missing = [n for n in names_in_file if n and not college_map.get(n)]
    if missing:
        print(f"  ⚠ Names with no college info ({len(missing)}):")
        for n in missing:
            print(f"    - {n}")

    # ── 3. Build new workbook ───────────────────────────────────────────
    print("\nBuilding output xlsx...")
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = ws.title

    # New headers: A(姓名), B(学号), C(学院), D(部门), E(身份), F(职能组), G(评优分组)
    new_headers = [
        headers[0],   # 姓名
        headers[1],   # 学号
        "学院",        # NEW
        headers[2],   # 部门
        headers[3],   # 身份
        headers[4],   # 职能组
        headers[5],   # 评优分组
    ]

    # Style for header
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write header
    for col_idx, h in enumerate(new_headers, 1):
        cell = new_ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data rows: insert college between B (index 1) and C (index 2)
    for row_idx, row in enumerate(rows, 2):
        name = str(row[0]).strip() if row[0] else ""
        college = college_map.get(name, "")

        new_row = [
            row[0],           # A: 姓名
            row[1],           # B: 学号
            college,          # C: 学院 (NEW)
            row[2],           # D: 部门
            row[3],           # E: 身份
            row[4],           # F: 职能组
            row[5],           # G: 评优分组
        ]

        for col_idx, val in enumerate(new_row, 1):
            cell = new_ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 3:  # 学院 column
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── 4. Adjust column widths ─────────────────────────────────────────
    col_widths = {
        "A": 10,   # 姓名
        "B": 18,   # 学号
        "C": 18,   # 学院
        "D": 28,   # 部门
        "E": 18,   # 身份
        "F": 18,   # 职能组
        "G": 45,   # 评优分组
    }
    for col_letter, width in col_widths.items():
        new_ws.column_dimensions[col_letter].width = width

    # Freeze top row
    new_ws.freeze_panes = "A2"

    # ── 5. Save ─────────────────────────────────────────────────────────
    new_wb.save(XLSX_OUT)
    print(f"\n✓ Saved to: {XLSX_OUT}")
    print(f"  {len(rows)} data rows + 1 header row")
    empty_count = sum(1 for n in names_in_file if n and not college_map.get(n, ""))
    filled_count = sum(1 for n in names_in_file if n and college_map.get(n, ""))
    print(f"  College filled: {filled_count}, Empty: {empty_count}")


if __name__ == "__main__":
    main()
